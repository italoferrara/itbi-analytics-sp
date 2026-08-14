from __future__ import annotations

import hashlib
import json
import math
import os
import queue
import re
import sqlite3
import sys
import threading
import traceback
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Callable, Iterable, Optional
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


CORE_API_VERSION = 202
CORE_VERSION = "WEB-0.2.2"

APP_TITLE = "FERRAMENTA DE BUSCA DE TRANSAÇÕES IMOBILIÁRIAS – ITBI SP"
SOURCE_PAGE = "https://prefeitura.sp.gov.br/web/fazenda/w/acesso_a_informacao/31501"
MONTHS = {
    "JAN": 1,
    "FEV": 2,
    "MAR": 3,
    "ABR": 4,
    "MAI": 5,
    "JUN": 6,
    "JUL": 7,
    "AGO": 8,
    "SET": 9,
    "OUT": 10,
    "NOV": 11,
    "DEZ": 12,
}
MONTH_NAMES = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Março",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro",
}
MONTH_SHEET_RE = re.compile(r"^(JAN|FEV|MAR|ABR|MAI|JUN|JUL|AGO|SET|OUT|NOV|DEZ)[-_ ]?(\d{4})$", re.I)
YEAR_RE = re.compile(r"\b(20(?:0[6-9]|1\d|2\d))\b")
CHUNK_SIZE = 1024 * 1024
BATCH_SIZE = 5000
SOURCE_COLUMN_COUNT = 28

CANONICAL_HEADERS = [
    "N° do Cadastro (SQL)",
    "Nome do Logradouro",
    "Número",
    "Complemento",
    "Bairro",
    "Referência",
    "CEP",
    "Natureza de Transação",
    "Valor de Transação (declarado pelo contribuinte)",
    "Data de Transação",
    "Valor Venal de Referência",
    "Proporção Transmitida (%)",
    "Valor Venal de Referência (proporcional)",
    "Base de Cálculo adotada",
    "Tipo de Financiamento",
    "Valor Financiado",
    "Cartório de Registro",
    "Matrícula do Imóvel",
    "Situação do SQL",
    "Área do Terreno (m2)",
    "Testada (m)",
    "Fração Ideal",
    "Área Construída (m2)",
    "Uso (IPTU)",
    "Descrição do uso (IPTU)",
    "Padrão (IPTU)",
    "Descrição do padrão (IPTU)",
    "ACC (IPTU)",
]

PREFIX_ALIASES = {
    "RUA": "R",
    "R": "R",
    "AVENIDA": "AV",
    "AV": "AV",
    "ALAMEDA": "AL",
    "AL": "AL",
    "PRACA": "PC",
    "PCA": "PC",
    "PC": "PC",
    "TRAVESSA": "TV",
    "TV": "TV",
    "ESTRADA": "EST",
    "EST": "EST",
    "RODOVIA": "ROD",
    "ROD": "ROD",
    "LARGO": "LGO",
    "LGO": "LGO",
    "VIELA": "VLA",
    "VLA": "VLA",
    "VIADUTO": "VD",
    "VD": "VD",
}
CANONICAL_PREFIXES = set(PREFIX_ALIASES.values())


@dataclass(frozen=True)
class SearchHit:
    year: int
    month: int
    file_path: str
    sheet_name: str
    row_number: int
    street: str
    number: str


@dataclass(frozen=True)
class SuggestionItem:
    display: str
    value: str
    key: str = ""
    neighborhoods: str = ""


def app_data_dir() -> Path:
    local_app_data = os.environ.get("LOCALAPPDATA")
    if local_app_data:
        base = Path(local_app_data)
    else:
        base = Path.home() / ".local" / "share"
    path = base / "Buscador_ITBI_SP"
    path.mkdir(parents=True, exist_ok=True)
    return path


BASE_DIR = app_data_dir()
DATA_DIR = BASE_DIR / "arquivos_xlsx"
DB_PATH = BASE_DIR / "indice_itbi.sqlite"
MANIFEST_PATH = BASE_DIR / "manifest.json"
CONFIG_PATH = BASE_DIR / "config.json"
LOG_PATH = BASE_DIR / "buscador_itbi.log"
DATA_DIR.mkdir(parents=True, exist_ok=True)


def log_exception(context: str) -> None:
    with LOG_PATH.open("a", encoding="utf-8") as handle:
        handle.write(f"\n[{datetime.now().isoformat(timespec='seconds')}] {context}\n")
        handle.write(traceback.format_exc())
        handle.write("\n")


def strip_accents(text: str) -> str:
    return "".join(
        char
        for char in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(char)
    )


def normalize_street(value: object) -> str:
    if value is None:
        return ""
    text = strip_accents(str(value)).upper().strip()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return ""
    parts = text.split()
    if parts[0] in PREFIX_ALIASES:
        parts[0] = PREFIX_ALIASES[parts[0]]
    return " ".join(parts)


def street_without_prefix(normalized: str) -> str:
    parts = normalized.split()
    if parts and parts[0] in CANONICAL_PREFIXES:
        return " ".join(parts[1:])
    return normalized


def input_has_prefix(normalized: str) -> bool:
    parts = normalized.split()
    return bool(parts and parts[0] in CANONICAL_PREFIXES)


def normalize_number(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return ("%.10f" % value).rstrip("0").rstrip(".")
    text = strip_accents(str(value)).upper().strip()
    text = text.replace(".", "").replace(" ", "")
    match = re.search(r"\d+", text)
    if match:
        return str(int(match.group(0)))
    return re.sub(r"[^A-Z0-9]", "", text)


def hash_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(CHUNK_SIZE), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_json(path: Path, default: dict) -> dict:
    try:
        with path.open("r", encoding="utf-8") as handle:
            value = json.load(handle)
            return value if isinstance(value, dict) else default
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return default


def save_json(path: Path, value: dict) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as handle:
        json.dump(value, handle, ensure_ascii=False, indent=2)
    tmp.replace(path)


def database_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH, timeout=60)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA temp_store=MEMORY")
    return conn


def initialize_database() -> None:
    with database_connection() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY,
                ano INTEGER NOT NULL,
                mes INTEGER NOT NULL,
                arquivo TEXT NOT NULL,
                aba TEXT NOT NULL,
                linha INTEGER NOT NULL,
                logradouro TEXT NOT NULL,
                numero TEXT NOT NULL,
                bairro TEXT NOT NULL DEFAULT '',
                logradouro_norm TEXT NOT NULL,
                logradouro_base TEXT NOT NULL,
                numero_norm TEXT NOT NULL,
                bairro_norm TEXT NOT NULL DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS indexed_files (
                ano INTEGER PRIMARY KEY,
                arquivo TEXT NOT NULL,
                sha256 TEXT NOT NULL,
                indexed_at TEXT NOT NULL,
                row_count INTEGER NOT NULL
            );
            CREATE TABLE IF NOT EXISTS street_catalog (
                logradouro_norm TEXT PRIMARY KEY,
                logradouro_base TEXT NOT NULL,
                logradouro TEXT NOT NULL,
                bairros TEXT NOT NULL,
                bairros_norm TEXT NOT NULL,
                record_count INTEGER NOT NULL
            );
            """
        )

        columns = {row[1] for row in conn.execute("PRAGMA table_info(records)")}
        migrated = False
        if "bairro" not in columns:
            conn.execute("ALTER TABLE records ADD COLUMN bairro TEXT NOT NULL DEFAULT ''")
            migrated = True
        if "bairro_norm" not in columns:
            conn.execute("ALTER TABLE records ADD COLUMN bairro_norm TEXT NOT NULL DEFAULT ''")
            migrated = True

        conn.executescript(
            """
            CREATE INDEX IF NOT EXISTS idx_records_exact
                ON records (logradouro_norm, numero_norm);
            CREATE INDEX IF NOT EXISTS idx_records_base
                ON records (logradouro_base, numero_norm);
            CREATE INDEX IF NOT EXISTS idx_records_address
                ON records (logradouro_norm, bairro_norm, numero_norm);
            CREATE INDEX IF NOT EXISTS idx_records_period
                ON records (ano DESC, mes DESC);
            CREATE INDEX IF NOT EXISTS idx_records_year
                ON records (ano);
            CREATE INDEX IF NOT EXISTS idx_street_catalog_base
                ON street_catalog (logradouro_base);
            """
        )

        # A versão 1.3 precisa reler as planilhas uma única vez para capturar o Bairro.
        # Os arquivos XLSX já baixados são reaproveitados; apenas o índice é reconstruído.
        if migrated:
            conn.execute("DELETE FROM records")
            conn.execute("DELETE FROM indexed_files")
            conn.execute("DELETE FROM street_catalog")
        conn.commit()


def extract_excel_links(session: requests.Session) -> dict[int, str]:
    response = session.get(SOURCE_PAGE, timeout=60)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")
    links: dict[int, str] = {}

    for anchor in soup.find_all("a", href=True):
        anchor_text = " ".join(anchor.get_text(" ", strip=True).split()).lower()
        href = anchor.get("href", "").strip()
        href_lower = href.lower()
        looks_excel = (
            "excel" in anchor_text
            or "xlsx" in anchor_text
            or ".xlsx" in href_lower
            or "xlsx" in href_lower
        )
        if not looks_excel:
            continue

        context = ""
        parent = anchor.find_parent(["li", "p", "div"])
        if parent is not None:
            context = parent.get_text(" ", strip=True)
        if not context:
            context = anchor.parent.get_text(" ", strip=True) if anchor.parent else anchor_text
        match = YEAR_RE.search(context)
        if match:
            year = int(match.group(1))
            links[year] = urljoin(response.url, href)

    if not links:
        # Fallback defensivo caso o HTML seja reorganizado: procura o ano nas proximidades do link.
        html = response.text
        for match in re.finditer(r"(?is)(20(?:0[6-9]|1\d|2\d)).{0,500}?href=[\"']([^\"']*(?:xlsx|XLSX)[^\"']*)[\"']", html):
            links[int(match.group(1))] = urljoin(response.url, match.group(2))

    expected_years = set(range(2006, datetime.now().year + 1))
    available = {year for year in links if year in expected_years}
    if not available:
        raise RuntimeError("Nenhum arquivo Excel anual foi localizado na página da Prefeitura.")

    return {year: links[year] for year in sorted(available)}


def download_file(
    session: requests.Session,
    year: int,
    url: str,
    destination: Path,
    progress: Callable[[str, Optional[float]], None],
) -> tuple[str, int]:
    tmp = destination.with_suffix(".download")
    response = session.get(url, stream=True, timeout=(30, 180), allow_redirects=True)
    response.raise_for_status()
    total = int(response.headers.get("content-length") or 0)
    written = 0
    digest = hashlib.sha256()

    with tmp.open("wb") as handle:
        for chunk in response.iter_content(CHUNK_SIZE):
            if not chunk:
                continue
            handle.write(chunk)
            digest.update(chunk)
            written += len(chunk)
            fraction = written / total if total else None
            progress(f"Baixando {year}: {written / (1024 * 1024):,.1f} MB", fraction)

    if written < 100_000:
        tmp.unlink(missing_ok=True)
        raise RuntimeError(f"O arquivo de {year} parece inválido ou incompleto.")

    tmp.replace(destination)
    return digest.hexdigest(), written


def workbook_month_sheets(workbook, expected_year: int) -> list[tuple[str, int]]:
    result: list[tuple[str, int]] = []
    for sheet_name in workbook.sheetnames:
        normalized = strip_accents(sheet_name).upper().strip()
        match = MONTH_SHEET_RE.match(normalized)
        if not match:
            continue
        month_code, year_text = match.groups()
        year = int(year_text)
        if year != expected_year:
            continue
        result.append((sheet_name, MONTHS[month_code.upper()]))
    return sorted(result, key=lambda item: item[1])


def index_year(
    year: int,
    file_path: Path,
    sha256: str,
    progress: Callable[[str, Optional[float]], None],
) -> int:
    progress(f"Abrindo o arquivo de {year} para indexação...", None)
    workbook = load_workbook(file_path, read_only=True, data_only=True, keep_links=False)
    month_sheets = workbook_month_sheets(workbook, year)
    if not month_sheets:
        workbook.close()
        raise RuntimeError(f"Nenhuma aba mensal válida foi encontrada no arquivo de {year}.")

    total_rows = 0
    with database_connection() as conn:
        conn.execute("DELETE FROM records WHERE ano = ?", (year,))
        conn.execute("DELETE FROM indexed_files WHERE ano = ?", (year,))

        for sheet_index, (sheet_name, month) in enumerate(month_sheets, start=1):
            ws = workbook[sheet_name]
            batch: list[tuple] = []
            rows_seen = 0
            for row_number, values in enumerate(
                ws.iter_rows(min_row=2, min_col=2, max_col=5, values_only=True),
                start=2,
            ):
                street, number, _complement, neighborhood = values
                if street is None and number is None:
                    continue
                street_text = "" if street is None else str(street).strip()
                number_text = "" if number is None else str(number).strip()
                neighborhood_text = "" if neighborhood is None else str(neighborhood).strip()
                street_norm = normalize_street(street_text)
                number_norm = normalize_number(number)
                if not street_norm or not number_norm:
                    continue
                batch.append(
                    (
                        year,
                        month,
                        str(file_path),
                        sheet_name,
                        row_number,
                        street_text,
                        number_text,
                        neighborhood_text,
                        street_norm,
                        street_without_prefix(street_norm),
                        number_norm,
                        normalize_street(neighborhood_text),
                    )
                )
                rows_seen += 1
                if len(batch) >= BATCH_SIZE:
                    conn.executemany(
                        """
                        INSERT INTO records
                        (ano, mes, arquivo, aba, linha, logradouro, numero, bairro,
                         logradouro_norm, logradouro_base, numero_norm, bairro_norm)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        batch,
                    )
                    batch.clear()
                    progress(
                        f"Indexando {sheet_name}: {rows_seen:,} linhas",
                        (sheet_index - 1) / len(month_sheets),
                    )
            if batch:
                conn.executemany(
                    """
                    INSERT INTO records
                    (ano, mes, arquivo, aba, linha, logradouro, numero, bairro,
                     logradouro_norm, logradouro_base, numero_norm, bairro_norm)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    batch,
                )
            total_rows += rows_seen
            progress(
                f"Indexado {sheet_name}: {rows_seen:,} registros",
                sheet_index / len(month_sheets),
            )

        conn.execute(
            """
            INSERT INTO indexed_files (ano, arquivo, sha256, indexed_at, row_count)
            VALUES (?, ?, ?, ?, ?)
            """,
            (year, str(file_path), sha256, datetime.now().isoformat(timespec="seconds"), total_rows),
        )
        conn.commit()

    workbook.close()
    return total_rows


def update_database(
    force_all: bool,
    progress: Callable[[str, Optional[float]], None],
    selected_years: Optional[Iterable[int]] = None,
) -> dict:
    initialize_database()
    manifest = load_json(MANIFEST_PATH, {})
    manifest.setdefault("years", {})
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 Chrome/124 Safari/537.36"
        )
    }
    session = requests.Session()
    session.headers.update(headers)
    all_links = extract_excel_links(session)
    source_current_year = max(all_links)

    if selected_years is None:
        links = all_links
    else:
        requested = sorted({int(year) for year in selected_years})
        links = {year: all_links[year] for year in requested if year in all_links}
        missing = [year for year in requested if year not in all_links]
        if missing:
            raise RuntimeError(
                "Os seguintes anos não estão disponíveis na página da Prefeitura: "
                + ", ".join(str(year) for year in missing)
            )
        if not links:
            raise RuntimeError("Selecione pelo menos um ano para atualizar.")

    summary = {
        "requested": list(links),
        "downloaded": [],
        "indexed": [],
        "skipped": [],
        "errors": [],
    }

    for position, (year, url) in enumerate(links.items(), start=1):
        progress(f"Preparando {year} ({position}/{len(links)})...", 0.0)
        destination = DATA_DIR / f"itbi_sp_{year}.xlsx"
        year_meta = manifest["years"].get(str(year), {})

        must_download = force_all or not destination.exists() or year == source_current_year
        downloaded_hash: Optional[str] = None
        downloaded_size: Optional[int] = None

        if must_download:
            try:
                previous_hash = year_meta.get("sha256")
                downloaded_hash, downloaded_size = download_file(
                    session, year, url, destination, progress
                )
                summary["downloaded"].append(year)
                changed = force_all or downloaded_hash != previous_hash
            except Exception as exc:
                if destination.exists():
                    progress(
                        f"Não foi possível baixar {year}; usando o arquivo local existente.",
                        None,
                    )
                    downloaded_hash = hash_file(destination)
                    downloaded_size = destination.stat().st_size
                    changed = False
                    summary["errors"].append(f"{year}: {exc}")
                else:
                    raise
        else:
            downloaded_hash = year_meta.get("sha256") or hash_file(destination)
            downloaded_size = destination.stat().st_size
            changed = False

        with database_connection() as conn:
            indexed = conn.execute(
                "SELECT sha256 FROM indexed_files WHERE ano = ?", (year,)
            ).fetchone()
        indexed_hash = indexed[0] if indexed else None

        needs_index = (
            force_all
            or indexed_hash is None
            or indexed_hash != downloaded_hash
            or changed
        )
        if needs_index:
            rows = index_year(year, destination, downloaded_hash, progress)
            summary["indexed"].append((year, rows))
        else:
            summary["skipped"].append(year)

        manifest["years"][str(year)] = {
            "url": url,
            "file": str(destination),
            "sha256": downloaded_hash,
            "size": downloaded_size,
            "checked_at": datetime.now().isoformat(timespec="seconds"),
        }
        save_json(MANIFEST_PATH, manifest)

    with database_connection() as conn:
        catalog_count = conn.execute("SELECT COUNT(*) FROM street_catalog").fetchone()[0]
    if summary["indexed"] or not catalog_count:
        summary["catalog_streets"] = rebuild_street_catalog(progress)
    else:
        summary["catalog_streets"] = int(catalog_count)

    manifest["last_update"] = datetime.now().isoformat(timespec="seconds")
    manifest["source_page"] = SOURCE_PAGE
    save_json(MANIFEST_PATH, manifest)
    return summary


def rebuild_street_catalog(
    progress: Optional[Callable[[str, Optional[float]], None]] = None,
) -> int:
    if progress:
        progress("Atualizando catálogo inteligente de ruas, bairros e números...", None)
    with database_connection() as conn:
        conn.execute("DELETE FROM street_catalog")
        grouped = conn.execute(
            """
            SELECT
                logradouro_norm,
                MIN(logradouro_base) AS logradouro_base,
                MIN(logradouro) AS logradouro,
                bairro_norm,
                MIN(bairro) AS bairro,
                COUNT(*) AS n
            FROM records
            GROUP BY logradouro_norm, bairro_norm
            """
        ).fetchall()

        catalog: dict[str, dict[str, object]] = {}
        for street_norm, street_base, street, neighborhood_norm, neighborhood, count in grouped:
            item = catalog.setdefault(
                str(street_norm),
                {
                    "base": str(street_base),
                    "street": str(street),
                    "count": 0,
                    "neighborhoods": [],
                },
            )
            item["count"] = int(item["count"]) + int(count)
            if neighborhood_norm and neighborhood:
                item["neighborhoods"].append((str(neighborhood), int(count)))

        batch: list[tuple[object, ...]] = []
        for street_norm, item in catalog.items():
            neighborhood_pairs = sorted(
                item["neighborhoods"],
                key=lambda pair: (-pair[1], normalize_street(pair[0])),
            )
            neighborhoods = ", ".join(name for name, _count in neighborhood_pairs[:12])
            batch.append(
                (
                    street_norm,
                    item["base"],
                    item["street"],
                    neighborhoods,
                    normalize_street(neighborhoods),
                    item["count"],
                )
            )

        conn.executemany(
            """
            INSERT INTO street_catalog
            (logradouro_norm, logradouro_base, logradouro, bairros, bairros_norm, record_count)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            batch,
        )
        conn.commit()
    return len(batch)


def database_status() -> tuple[int, int, Optional[str], Optional[int], Optional[int]]:
    initialize_database()
    with database_connection() as conn:
        years, rows = conn.execute(
            "SELECT COUNT(*), COALESCE(SUM(row_count), 0) FROM indexed_files"
        ).fetchone()
        latest = conn.execute(
            "SELECT ano, mes FROM records ORDER BY ano DESC, mes DESC LIMIT 1"
        ).fetchone()
    manifest = load_json(MANIFEST_PATH, {})
    latest_year = int(latest[0]) if latest else None
    latest_month = int(latest[1]) if latest else None
    return int(years), int(rows), manifest.get("last_update"), latest_year, latest_month


def indexed_years() -> list[int]:
    initialize_database()
    with database_connection() as conn:
        rows = conn.execute("SELECT ano FROM indexed_files ORDER BY ano").fetchall()
    return [int(row[0]) for row in rows]


def list_streets(limit: Optional[int] = None) -> list[SuggestionItem]:
    """Return the street catalog for a native searchable Streamlit selectbox."""
    initialize_database()
    query = """
        SELECT logradouro, logradouro_norm, bairros, record_count
        FROM street_catalog
        ORDER BY logradouro
    """
    params: tuple[object, ...] = ()
    if limit is not None:
        query += " LIMIT ?"
        params = (int(limit),)
    with database_connection() as conn:
        rows = conn.execute(query, params).fetchall()

    suggestions: list[SuggestionItem] = []
    for street, street_norm, neighborhoods, _count in rows:
        cleaned = _clean_neighborhoods(neighborhoods)
        display = str(street)
        if cleaned:
            display += f"  •  {cleaned}"
        suggestions.append(
            SuggestionItem(
                display=display,
                value=str(street),
                key=str(street_norm),
                neighborhoods=cleaned,
            )
        )
    return suggestions


def _clean_neighborhoods(value: object, limit: int = 4) -> str:
    parts: list[str] = []
    seen: set[str] = set()
    for raw in str(value or "").split(","):
        item = raw.strip()
        norm = normalize_street(item)
        if not item or not norm or norm in seen:
            continue
        seen.add(norm)
        parts.append(item)
    parts.sort(key=lambda item: normalize_street(item))
    if len(parts) > limit:
        return ", ".join(parts[:limit]) + f" +{len(parts) - limit}"
    return ", ".join(parts)


def suggest_streets(text: str, limit: int = 30) -> list[SuggestionItem]:
    normalized = normalize_street(text)
    base = street_without_prefix(normalized)
    tokens = [token for token in base.split() if token]
    if sum(len(token) for token in tokens) < 2 or not DB_PATH.exists():
        return []

    where = " AND ".join("logradouro_base LIKE ?" for _ in tokens)
    params: list[object] = [f"%{token}%" for token in tokens]
    prefix = f"{base}%"
    params.extend([prefix, limit])
    with database_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT logradouro, logradouro_norm, bairros, record_count
            FROM street_catalog
            WHERE {where}
            ORDER BY CASE WHEN logradouro_base LIKE ? THEN 0 ELSE 1 END,
                     record_count DESC, logradouro
            LIMIT ?
            """,
            params,
        ).fetchall()

    suggestions: list[SuggestionItem] = []
    for street, street_norm, neighborhoods, _count in rows:
        cleaned = _clean_neighborhoods(neighborhoods)
        display = str(street)
        if cleaned:
            display += f"   •   Bairro(s): {cleaned}"
        suggestions.append(
            SuggestionItem(
                display=display,
                value=str(street),
                key=str(street_norm),
                neighborhoods=cleaned,
            )
        )
    return suggestions


def suggest_numbers(
    street: str,
    typed_number: str = "",
    selected_street_norm: str = "",
    limit: int = 100,
) -> list[SuggestionItem]:
    street_norm = selected_street_norm or normalize_street(street)
    base_norm = street_without_prefix(street_norm)
    if not street_norm or not DB_PATH.exists():
        return []

    conditions: list[str] = []
    params: list[object] = []
    if selected_street_norm or input_has_prefix(street_norm):
        conditions.append("logradouro_norm = ?")
        params.append(street_norm)
    else:
        conditions.append("logradouro_base = ?")
        params.append(base_norm)

    number_filter = normalize_number(typed_number)
    if typed_number.strip() and number_filter:
        conditions.append("numero_norm LIKE ?")
        params.append(f"{number_filter}%")

    with database_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT
                numero_norm,
                MIN(numero) AS numero,
                bairro_norm,
                MIN(bairro) AS bairro,
                COUNT(*) AS n
            FROM records
            WHERE {' AND '.join(conditions)}
            GROUP BY numero_norm, bairro_norm
            """,
            params,
        ).fetchall()

    grouped_numbers: dict[str, dict[str, object]] = {}
    for number_norm, number, neighborhood_norm, neighborhood, count in rows:
        item = grouped_numbers.setdefault(
            str(number_norm),
            {"number": str(number), "neighborhoods": [], "count": 0},
        )
        item["count"] = int(item["count"]) + int(count)
        if neighborhood_norm and neighborhood:
            item["neighborhoods"].append((str(neighborhood), int(count)))

    def number_sort_key(item: tuple[str, dict[str, object]]):
        normalized_number = item[0]
        if normalized_number.isdigit():
            return (0, int(normalized_number), normalized_number)
        return (1, 0, normalized_number)

    suggestions: list[SuggestionItem] = []
    for number_norm, item in sorted(grouped_numbers.items(), key=number_sort_key)[:limit]:
        neighborhood_pairs = sorted(
            item["neighborhoods"],
            key=lambda pair: (-pair[1], normalize_street(pair[0])),
        )
        cleaned = _clean_neighborhoods(
            ", ".join(name for name, _count in neighborhood_pairs),
            limit=4,
        )
        display = str(item["number"])
        if cleaned:
            display += f"   •   Bairro(s): {cleaned}"
        suggestions.append(
            SuggestionItem(
                display=display,
                value=str(item["number"]),
                key=number_norm,
                neighborhoods=cleaned,
            )
        )
    return suggestions


def find_hits(street: str, number: str, mode: str = "exact") -> list[SearchHit]:
    street_norm = normalize_street(street)
    base_norm = street_without_prefix(street_norm)
    number_norm = normalize_number(number)
    if not street_norm or not number_norm:
        return []

    with database_connection() as conn:
        if mode == "contains":
            query = """
                SELECT ano, mes, arquivo, aba, linha, logradouro, numero
                FROM records
                WHERE logradouro_base LIKE ? AND numero_norm = ?
                ORDER BY ano, mes, linha
            """
            params = (f"%{base_norm}%", number_norm)
        elif input_has_prefix(street_norm):
            query = """
                SELECT ano, mes, arquivo, aba, linha, logradouro, numero
                FROM records
                WHERE logradouro_norm = ? AND numero_norm = ?
                ORDER BY ano, mes, linha
            """
            params = (street_norm, number_norm)
        else:
            query = """
                SELECT ano, mes, arquivo, aba, linha, logradouro, numero
                FROM records
                WHERE logradouro_base = ? AND numero_norm = ?
                ORDER BY ano, mes, linha
            """
            params = (base_norm, number_norm)

        rows = conn.execute(query, params).fetchall()

    return [SearchHit(*row) for row in rows]


def excel_safe_value(value: object) -> object:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return value
    if value is None:
        return None
    return value


def load_source_rows(
    hits: list[SearchHit],
    progress: Callable[[str, Optional[float]], None],
) -> list[list[object]]:
    grouped: dict[str, dict[str, set[int]]] = defaultdict(lambda: defaultdict(set))
    hit_lookup: dict[tuple[str, str, int], SearchHit] = {}
    for hit in hits:
        grouped[hit.file_path][hit.sheet_name].add(hit.row_number)
        hit_lookup[(hit.file_path, hit.sheet_name, hit.row_number)] = hit

    results: list[list[object]] = []
    total_sheets = sum(len(sheets) for sheets in grouped.values())
    sheet_counter = 0

    for file_path, sheets in grouped.items():
        workbook = load_workbook(file_path, read_only=True, data_only=True, keep_links=False)
        try:
            for sheet_name, row_numbers in sheets.items():
                sheet_counter += 1
                progress(
                    f"Lendo {Path(file_path).name} / {sheet_name}",
                    sheet_counter / max(total_sheets, 1),
                )
                ws = workbook[sheet_name]
                minimum = min(row_numbers)
                maximum = max(row_numbers)
                for row_no, values in enumerate(
                    ws.iter_rows(
                        min_row=minimum,
                        max_row=maximum,
                        min_col=1,
                        max_col=SOURCE_COLUMN_COUNT,
                        values_only=True,
                    ),
                    start=minimum,
                ):
                    if row_no not in row_numbers:
                        continue
                    hit = hit_lookup[(file_path, sheet_name, row_no)]
                    row = [
                        hit.year,
                        MONTH_NAMES.get(hit.month, str(hit.month)),
                        sheet_name,
                        row_no,
                    ]
                    row.extend(excel_safe_value(value) for value in values)
                    results.append(row)
        finally:
            workbook.close()

    results.sort(key=lambda row: (row[0], MONTHS.get(str(row[1])[:3].upper(), 99), row[3]))
    return results


def set_column_widths(ws, widths: dict[int, float]) -> None:
    for index, width in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = width



APP_VERSION = "WEB-0.2.2"
OUTPUT_OFFSET = 4
TRANSACTION_VALUE_INDEX = OUTPUT_OFFSET + 8
TRANSACTION_DATE_INDEX = OUTPUT_OFFSET + 9
BUILT_AREA_INDEX = OUTPUT_OFFSET + 22
COMPLEMENT_INDEX = OUTPUT_OFFSET + 3
NATURE_INDEX = OUTPUT_OFFSET + 7
BAIRRO_INDEX = OUTPUT_OFFSET + 4
USE_DESCRIPTION_INDEX = OUTPUT_OFFSET + 24


@dataclass
class CellStats:
    count: int = 0
    total: float = 0.0
    minimum: Optional[float] = None
    maximum: Optional[float] = None

    def add(self, value: float) -> None:
        self.count += 1
        self.total += value
        self.minimum = value if self.minimum is None else min(self.minimum, value)
        self.maximum = value if self.maximum is None else max(self.maximum, value)

    @property
    def average(self) -> Optional[float]:
        return self.total / self.count if self.count else None


@dataclass
class AnalyticsResult:
    years: list[int]
    area_keys: list[Optional[int]]
    cells: dict[tuple[Optional[int], int], CellStats]
    total_records: int
    purchase_sale_records: int
    priced_records: int
    overall_average: Optional[float]
    overall_minimum: Optional[float]
    overall_maximum: Optional[float]
    years_with_sales: int
    known_typologies: int
    minimum_area: Optional[int]
    maximum_area: Optional[int]


def coerce_number(value: object) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None

    text = str(value).strip()
    if not text:
        return None
    text = text.replace("R$", "").replace("m²", "").replace("m2", "")
    text = text.replace("\xa0", "").replace(" ", "")
    text = re.sub(r"[^0-9,.-]", "", text)
    if not text or text in {"-", ",", "."}:
        return None

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")

    try:
        number = float(text)
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def area_group(value: object) -> Optional[int]:
    number = coerce_number(value)
    if number is None or number <= 0:
        return None
    try:
        return int(Decimal(str(number)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    except (InvalidOperation, ValueError):
        return None


def format_brl(value: Optional[float], decimals: int = 2) -> str:
    if value is None:
        return "—"
    formatted = f"{value:,.{decimals}f}"
    formatted = formatted.replace(",", "§").replace(".", ",").replace("§", ".")
    return f"R$ {formatted}"


def format_brl_compact(value: Optional[float]) -> str:
    if value is None:
        return ""
    absolute = abs(value)
    if absolute >= 1_000_000_000:
        return f"R$ {value / 1_000_000_000:.2f} bi".replace(".", ",")
    if absolute >= 1_000_000:
        return f"R$ {value / 1_000_000:.2f} mi".replace(".", ",")
    formatted = f"{value:,.0f}".replace(",", ".")
    return f"R$ {formatted}"


def format_area(area: Optional[int]) -> str:
    if area is None:
        return "Sem m² informado"
    return f"{area:,} m²".replace(",", ".")


def format_plain_number(value: Optional[float], decimals: int = 0) -> str:
    if value is None:
        return "—"
    formatted = f"{value:,.{decimals}f}"
    return formatted.replace(",", "§").replace(".", ",").replace("§", ".")


def is_purchase_and_sale(value: object) -> bool:
    normalized = normalize_street(value)
    return normalized == "1 COMPRA E VENDA"


def row_is_used_in_analytics(row: list[object]) -> bool:
    if len(row) <= max(TRANSACTION_VALUE_INDEX, NATURE_INDEX):
        return False
    price = coerce_number(row[TRANSACTION_VALUE_INDEX])
    return is_purchase_and_sale(row[NATURE_INDEX]) and price is not None and price > 0


def build_analytics(rows: list[list[object]]) -> AnalyticsResult:
    years_in_rows = [int(row[0]) for row in rows if row and isinstance(row[0], (int, float))]
    final_year = max([datetime.now().year, *years_in_rows], default=datetime.now().year)
    years = list(range(2006, final_year + 1))
    cells: dict[tuple[Optional[int], int], CellStats] = {}
    overall = CellStats()
    used_years: set[int] = set()
    known_areas: set[int] = set()
    saw_unknown_area = False
    purchase_sale_records = 0

    for row in rows:
        if len(row) <= max(TRANSACTION_VALUE_INDEX, BUILT_AREA_INDEX, NATURE_INDEX):
            continue
        if not is_purchase_and_sale(row[NATURE_INDEX]):
            continue
        purchase_sale_records += 1
        try:
            year = int(row[0])
        except (TypeError, ValueError):
            continue
        price = coerce_number(row[TRANSACTION_VALUE_INDEX])
        if price is None or price <= 0:
            continue
        area = area_group(row[BUILT_AREA_INDEX])
        if area is None:
            saw_unknown_area = True
        else:
            known_areas.add(area)
        stats = cells.setdefault((area, year), CellStats())
        stats.add(price)
        overall.add(price)
        used_years.add(year)

    area_keys: list[Optional[int]] = sorted(known_areas)
    if saw_unknown_area:
        area_keys.append(None)

    return AnalyticsResult(
        years=years,
        area_keys=area_keys,
        cells=cells,
        total_records=len(rows),
        purchase_sale_records=purchase_sale_records,
        priced_records=overall.count,
        overall_average=overall.average,
        overall_minimum=overall.minimum,
        overall_maximum=overall.maximum,
        years_with_sales=len(used_years),
        known_typologies=len(known_areas),
        minimum_area=min(known_areas) if known_areas else None,
        maximum_area=max(known_areas) if known_areas else None,
    )


def add_matrix_sheet(
    workbook: Workbook,
    title: str,
    sheet_name: str,
    analytics: AnalyticsResult,
    metric: str,
    colors: dict[str, str],
) -> None:
    ws = workbook.create_sheet(sheet_name)
    end_col = 1 + len(analytics.years)
    end_letter = get_column_letter(end_col)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, color=colors["white"], size=15)
    ws["A1"].fill = PatternFill("solid", fgColor=colors["navy"])
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws["A2"] = (
        "Tipologias agrupadas pela coluna Área Construída (m²), arredondada ao m² inteiro. "
        "Somente Natureza de Transação = 1. Compra e Venda e valores declarados positivos entram nos cálculos."
    )
    ws["A2"].font = Font(italic=True, color=colors["muted"], size=9)
    ws["A2"].alignment = Alignment(wrap_text=True)

    header_row = 4
    ws.cell(header_row, 1, "Tipologia por área")
    for col, year in enumerate(analytics.years, start=2):
        ws.cell(header_row, col, year)
    for cell in ws[header_row]:
        cell.fill = PatternFill("solid", fgColor=colors["blue"])
        cell.font = Font(bold=True, color=colors["white"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="medium", color=colors["navy"]))
    ws.row_dimensions[header_row].height = 30

    for row_index, area in enumerate(analytics.area_keys, start=header_row + 1):
        ws.cell(row_index, 1, format_area(area))
        ws.cell(row_index, 1).font = Font(bold=True, color=colors["navy"])
        for col, year in enumerate(analytics.years, start=2):
            stats = analytics.cells.get((area, year))
            if not stats:
                continue
            value = stats.average if metric == "average" else stats.count
            ws.cell(row_index, col, value)
            if metric == "average":
                ws.cell(row_index, col).number_format = 'R$ #,##0.00'
            else:
                ws.cell(row_index, col).number_format = '0'
        if row_index % 2 == 1:
            for cell in ws[row_index]:
                cell.fill = PatternFill("solid", fgColor=colors["light_gray"])

    ws.freeze_panes = "B5"
    ws.auto_filter.ref = f"A{header_row}:{end_letter}{max(header_row, header_row + len(analytics.area_keys))}"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    for col in range(2, end_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15


def export_results(
    destination: Path,
    query_street: str,
    query_number: str,
    mode: str,
    rows: list[list[object]],
    analytics: Optional[AnalyticsResult] = None,
) -> None:
    analytics = analytics or build_analytics(rows)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumo"

    colors = {
        "navy": "102A43",
        "blue": "2563EB",
        "cyan": "0EA5E9",
        "light_blue": "E8F1FF",
        "light_gray": "F4F7FB",
        "white": "FFFFFF",
        "muted": "52667A",
        "green": "0F9D76",
    }
    thin_gray = Side(style="thin", color="D8E2EC")

    summary.merge_cells("A1:L1")
    summary["A1"] = "ITBI ANALYTICS SP — RELATÓRIO DO ENDEREÇO"
    summary["A1"].font = Font(bold=True, color=colors["white"], size=16)
    summary["A1"].fill = PatternFill("solid", fgColor=colors["navy"])
    summary["A1"].alignment = Alignment(horizontal="left", vertical="center")
    summary.row_dimensions[1].height = 30

    summary.merge_cells("A2:L2")
    summary["A2"] = "Transações imobiliárias com recolhimento de ITBI — Município de São Paulo"
    summary["A2"].font = Font(color=colors["muted"], italic=True)

    query_data = [
        ("Logradouro pesquisado", query_street),
        ("Número pesquisado", query_number),
        ("Modo de busca", "Contém" if mode == "contains" else "Exato"),
        ("Data da consulta", datetime.now()),
        ("Fonte oficial", SOURCE_PAGE),
    ]
    for row_index, (label, value) in enumerate(query_data, start=4):
        summary.cell(row_index, 1, label)
        summary.cell(row_index, 2, value)
        summary.cell(row_index, 1).font = Font(bold=True, color=colors["navy"])
        summary.cell(row_index, 1).fill = PatternFill("solid", fgColor=colors["light_blue"])
        summary.cell(row_index, 1).border = Border(bottom=thin_gray)
        summary.cell(row_index, 2).border = Border(bottom=thin_gray)
    summary["B7"].number_format = "dd/mm/yyyy hh:mm"
    summary["B8"].hyperlink = SOURCE_PAGE
    summary["B8"].style = "Hyperlink"

    kpis = [
        ("Transações encontradas", analytics.total_records, "0"),
        ("Compras e vendas", analytics.purchase_sale_records, "0"),
        ("Vendas com preço válido", analytics.priced_records, "0"),
        ("Preço médio declarado", analytics.overall_average, 'R$ #,##0.00'),
        ("Tipologias por m²", analytics.known_typologies, "0"),
        ("Anos com transações", analytics.years_with_sales, "0"),
    ]
    summary.merge_cells("A10:L10")
    summary["A10"] = "CRITÉRIO ANALÍTICO: somente Natureza de Transação = 1. Compra e Venda"
    summary["A10"].font = Font(bold=True, color=colors["green"], size=9)
    for index, (label, value, number_format) in enumerate(kpis):
        start_col = 1 + index * 2
        summary.merge_cells(start_row=11, start_column=start_col, end_row=11, end_column=start_col + 1)
        summary.merge_cells(start_row=12, start_column=start_col, end_row=13, end_column=start_col + 1)
        label_cell = summary.cell(11, start_col)
        value_cell = summary.cell(12, start_col)
        label_cell.value = label
        value_cell.value = value
        label_cell.fill = PatternFill("solid", fgColor=colors["blue"])
        label_cell.font = Font(bold=True, color=colors["white"], size=9)
        label_cell.alignment = Alignment(horizontal="center")
        value_cell.fill = PatternFill("solid", fgColor=colors["light_blue"])
        value_cell.font = Font(bold=True, color=colors["navy"], size=15)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = number_format
        summary.cell(12, start_col + 1).fill = PatternFill("solid", fgColor=colors["light_blue"])
        summary.cell(13, start_col).fill = PatternFill("solid", fgColor=colors["light_blue"])
        summary.cell(13, start_col + 1).fill = PatternFill("solid", fgColor=colors["light_blue"])

    details = [
        ("Menor preço declarado", analytics.overall_minimum, 'R$ #,##0.00'),
        ("Maior preço declarado", analytics.overall_maximum, 'R$ #,##0.00'),
        ("Menor área construída", analytics.minimum_area, '0 "m²"'),
        ("Maior área construída", analytics.maximum_area, '0 "m²"'),
    ]
    for row_index, (label, value, number_format) in enumerate(details, start=16):
        summary.cell(row_index, 1, label)
        summary.cell(row_index, 2, value)
        summary.cell(row_index, 1).font = Font(bold=True, color=colors["navy"])
        summary.cell(row_index, 2).number_format = number_format

    summary.merge_cells("A22:L24")
    summary["A22"] = (
        "Nota metodológica: o preço utiliza exclusivamente a coluna 'Valor de Transação (declarado pelo contribuinte)'. "
        "Entram nos cálculos apenas registros cuja coluna 'Natureza de Transação' seja '1. Compra e Venda' e cujo valor declarado seja positivo. "
        "As tipologias utilizam exclusivamente a coluna 'Área Construída (m²)', arredondada ao inteiro mais próximo. "
        "A aba de transações detalhadas preserva todos os registros encontrados e o complemento integral, sem truncamento."
    )
    summary["A22"].alignment = Alignment(wrap_text=True, vertical="top")
    summary["A22"].font = Font(color=colors["muted"], size=9)
    summary.sheet_view.showGridLines = False
    set_column_widths(summary, {
        1: 28, 2: 24, 3: 20, 4: 20, 5: 20, 6: 20,
        7: 20, 8: 20, 9: 20, 10: 20, 11: 20, 12: 20,
    })

    add_matrix_sheet(
        workbook,
        "PREÇO MÉDIO DECLARADO DE COMPRA E VENDA POR ÁREA CONSTRUÍDA E ANO",
        "Preço médio por área",
        analytics,
        "average",
        colors,
    )
    add_matrix_sheet(
        workbook,
        "QUANTIDADE DE COMPRAS E VENDAS POR ÁREA CONSTRUÍDA E ANO",
        "Quantidade por área",
        analytics,
        "count",
        colors,
    )

    results_ws = workbook.create_sheet("Transações detalhadas")
    output_headers = ["Ano", "Mês", "Aba de origem", "Linha de origem"] + CANONICAL_HEADERS
    results_ws.append(output_headers)
    for row in rows:
        results_ws.append(row)

    header_fill = PatternFill("solid", fgColor=colors["blue"])
    header_font = Font(bold=True, color=colors["white"])
    for cell in results_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=colors["navy"]))
    results_ws.row_dimensions[1].height = 42
    results_ws.freeze_panes = "A2"
    results_ws.auto_filter.ref = results_ws.dimensions
    results_ws.sheet_view.showGridLines = False

    date_column = OUTPUT_OFFSET + 10
    currency_source_columns = [9, 11, 13, 14, 16]
    decimal_source_columns = [12, 20, 21, 22, 23]
    for row_index in range(2, results_ws.max_row + 1):
        results_ws.cell(row_index, date_column).number_format = "dd/mm/yyyy"
        for source_index in currency_source_columns:
            results_ws.cell(row_index, OUTPUT_OFFSET + source_index).number_format = 'R$ #,##0.00'
        for source_index in decimal_source_columns:
            results_ws.cell(row_index, OUTPUT_OFFSET + source_index).number_format = '#,##0.00'
        if row_index % 2 == 0:
            for cell in results_ws[row_index]:
                cell.fill = PatternFill("solid", fgColor=colors["light_gray"])

    if results_ws.max_row >= 2:
        table_ref = f"A1:{get_column_letter(results_ws.max_column)}{results_ws.max_row}"
        table = Table(displayName="TabelaTransacoesITBI", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        results_ws.add_table(table)

    widths = {
        1: 9, 2: 13, 3: 15, 4: 14, 5: 17, 6: 35, 7: 10, 8: 22,
        9: 22, 10: 22, 11: 12, 12: 34, 13: 20, 14: 15, 15: 20,
        16: 16, 17: 20, 18: 20, 19: 18, 20: 25, 21: 18, 22: 20,
        23: 18, 24: 18, 25: 18, 26: 18, 27: 16, 28: 20, 29: 35,
        30: 16, 31: 35, 32: 15,
    }
    set_column_widths(results_ws, widths)

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)


